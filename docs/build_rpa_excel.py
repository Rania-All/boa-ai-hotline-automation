# -*- coding: utf-8 -*-
"""Génère le fichier Excel RPA-Detail-Robot-Architecture-Etapes.xlsx"""
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installation de openpyxl...")
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Side(style='thin')
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
title_font = Font(bold=True, size=12)

def sheet_vue_ensemble(ws):
    ws.title = "1_Vue_ensemble"
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 70
    rows = [
        ["Élément", "Description"],
        ["Projet", "Automatisation Hotline (version Python)"],
        ["Partie 1 - Chatbot", "Chatbot utilisant un fichier Excel (questions / réponses FAQ). À partir de la question utilisateur, une classification est effectuée pour retrouver la réponse ou identifier un cas N1-RR."],
        ["Partie 2 - RPA", "Lorsque la classification indique N1-RR, le backend Python déclenche et suit les processus UiPath via l'Orchestrator."],
        ["Fichier Excel FAQ", "Contient les paires question / réponse ; utilisé par le chatbot pour les réponses directes et pour alimenter la classification (ex. N1-RR)."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_architecture(ws):
    ws.title = "2_Architecture"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60
    rows = [
        ["Composant", "Rôle"],
        ["Application Python (RAG)", "Backend : classification N1-RR via RAG/IA, vérification des données, appel des APIs UiPath (client HTTP), suivi des jobs, retour au client."],
        ["UiPath Orchestrator", "Plateforme centrale : authentification (OAuth2), envoi des jobs aux robots, suivi du statut."],
        ["UiPath Robots", "Exécution des processus RPA (traitement N1-RR, automatisation des tâches hotline)."],
        ["", ""],
        ["Configuration (Python)", "Variables d'environnement ou .env : UIPATH_BASE_URL, UIPATH_CLIENT_ID, UIPATH_CLIENT_SECRET, UIPATH_FOLDER_ID, UIPATH_RELEASE_KEY"],
        ["Client Orchestrator", "Module requests ou httpx : OAuth2, POST Start Jobs, GET Jobs pour le statut."],
        ["API exposée", "Endpoints /api/rpa/start et /api/rpa/status/{job_key} (ex. FastAPI/Flask)."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_etapes(ws):
    ws.title = "3_Etapes_RPA"
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 55
    rows = [
        ["Étape", "Titre", "Détail"],
        [1, "Classification N1-RR", "Chatbot + fichier Excel FAQ → classification. Si N1-RR : envoi vers RPA. Sinon réponse depuis la FAQ."],
        [2, "Vérification des données", "Contrôles côté backend avant appel Start Job (données requises présentes et valides)."],
        [3, "Authentification OAuth2", "POST vers https://cloud.uipath.com/identity_/connect/token ; client_id, client_secret, scope OR.Folders OR.Jobs ; récupération access_token."],
        [4, "Appel API Start Job", "POST /odata/Jobs/UiPath.Server.Configuration.OData.StartJobs ; headers Authorization Bearer, X-UIPATH-OrganizationUnitId ; body startInfo (ReleaseKey, InputArguments, etc.)."],
        [5, "InputArguments en JSON", "Données métier (action, payload) sérialisées en JSON (json.dumps) dans le corps de la requête Start Job."],
        [6, "Suivi Get Job Status", "GET OData Jobs avec $filter=Key eq '<job_key>' ; exposé via GET /api/rpa/status/{job_key}. Polling jusqu'à terminaison ou timeout."],
        [7, "Retour au client", "Retour immédiat job_key + état initial ; le client peut interroger /api/rpa/status/{job_key} pour le suivi."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=3):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_securite(ws):
    ws.title = "4_Securite_Conformite"
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 55
    rows = [
        ["Mesure", "Description"],
        ["Masquage des données sensibles", "Ex. numéro de carte masqué avant envoi ou affichage (côté métier / logs)."],
        ["Journalisation des actions", "Appels RPA (démarrage, statut) et erreurs loggés pour traçabilité et audit."],
        ["Gestion des erreurs et escalade", "En cas d'échec (token, API, timeout) : gestion d'erreur et escalade automatique (notification, repli manuel)."],
        ["Secrets", "client_id, client_secret dans variables d'environnement ou .env (non versionné)."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_kpis(ws):
    ws.title = "5_KPIs"
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 45
    rows = [
        ["KPI", "Objectif"],
        ["Taux de précision de classification", "> 85 % (module RAG/IA pour N1-RR)."],
        ["Taux d'automatisation N1-RR", "> 60 % des cas N1-RR traités par RPA."],
        ["Réduction du temps moyen de traitement", "Objectif global du projet."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sheet_planning(ws):
    ws.title = "6_Planning"
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 60
    rows = [
        ["Période", "Activités / Livrables"],
        ["Mois 1", "Formation UiPath Academy ; conception architecture Python (RAG + API) ; spécifications fonctionnelles et techniques (dont RPA)."],
        ["Mois 2", "Développement chaîne RAG ; API REST (FastAPI/Flask) ; moteur de décision ; préparation données et classification pour RPA. Chatbot + fichier Excel FAQ."],
        ["Mois 3 (S9-S12)", "Intégration UiPath : service d'authentification Orchestrator (OAuth2) ; déclenchement job via API (Start Job + InputArguments) ; tests avec robots existants ; gestion logs et erreurs. Livrable : Prototype end-to-end fonctionnel."],
        ["Mois 4", "Tests utilisateurs (hotline) ; optimisation performances ; documentation technique complète. Livrable : Solution prête pour pilote."],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            if i == 1:
                c.fill = header_fill
                c.font = header_font
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Feuille 1 par défaut
sheet_vue_ensemble(wb.active)
# Autres feuilles
sheet_architecture(wb.create_sheet())
sheet_etapes(wb.create_sheet())
sheet_securite(wb.create_sheet())
sheet_kpis(wb.create_sheet())
sheet_planning(wb.create_sheet())

out = os.path.join(os.path.dirname(__file__), "RPA-Detail-Robot-Architecture-Etapes.xlsx")
wb.save(out)
print("Fichier créé :", out)
